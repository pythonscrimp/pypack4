from selenium import webdriver as wd
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait as wdw
from selenium.webdriver.support import expected_conditions as ec
from selenium.common.exceptions import TimeoutException as te
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.firefox.firefox_binary import FirefoxBinary
from tinydb import TinyDB, Query
from flask import Flask, jsonify, request
import time
import datetime
from datetime import datetime
import openpyxl as XL
from openpyxl.workbook import Workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell
import os
from pprint import pprint

DB = TinyDB('Settings.json')
dbq = Query()



IPO = 'Do you want to consider PO number? '
IZI = 'Found COUPA invoices, do you want to search? '
IVN = 'Do you want to search the invoices with invoice # as well as Vendor name? '


usdat = []
nusdat = []
Nusdat = []
vendat = []
vendatE = []
unnum = []
Unnum = []
dunum = []
ponum = []
pdunum = []
invnu = []
manua = []


def conNcon(q): # Questions for Input File
	Q = input(q)
	if Q.upper() == 'Y':
		if q == IPO:
			print('Po number will be considered.')
			checkValidPO(inputData)
		elif q == IVN:
			print('Invoice will be searching with both invoice number and vendor name.')
			sivinav()
		elif q == IZI:
			print('Coupa invoice will search')
			Cfci()		
	elif Q.upper() == 'N':
		if q == IPO:
			print('Po number will not be considered.')
			DonotcheckValidPO(inputData)
		elif q == IVN:
			print('Invoice will start searching using invoice number only')
			sivin()
		elif q == IZI:
			print('Coupa invoice will not search')			
	else:
		if q == IPO:
			print('Please Select Yes or No, whereas Y as Yes and N as No')
			conNcon(IPO)
		if q == IVN:
			print('Please Select Yes or No, whereas Y as Yes and N as No')
			conNcon(IVN)
		if q == IZI:
			print('Please Select Yes or No, whereas Y as Yes and N as No')
			conNcon(IZI)			



inputFile = []
inputData = []

filePath = '1_Raw_Files\\Input'
Path = os.path.join(os.path.dirname(os.path.abspath(__name__)),filePath)
DrivePat = '\\IEDriverServer.exe'
DriverPath = os.path.join(os.path.dirname(os.path.abspath(__name__)),DrivePat)

def getInputFile():
	for Dir, SubDir, Files in os.walk(Path):
		for file in Files:
			if file.endswith('.xlsx'):
				inputFile.append(os.path.join(Path,file))

getInputFile()



def readInputFile(file):
	WB = XL.load_workbook(file[0],data_only=True)
	Sheet = WB[WB.sheetnames[0]]
	Val = str(input('Enter last row #: '))
	for data,Data in enumerate(Sheet['A2':'H'+Val]):
		inputData.append([])
		for value in Data:
			inputData[data].append(str(value.value))

readInputFile(inputFile)

# pprint(inputFile)

POtable = DB.table('poTable').get(doc_id=1)['valdPO']

ValidPO = []
InValidPO = []

def checkValidPO(PO):
	for Vpo,vPO in enumerate(PO):
		if (vPO[7][:3] in POtable) or (vPO[7][:4] in POtable) or (vPO[7][:2] in POtable):
			ValidPO.append(vPO)
		else:
			InValidPO.append(vPO)

def DonotcheckValidPO(PO):
	for Vpo,vPO in enumerate(PO):
		ValidPO.append(vPO)

conNcon(IPO)

def Drive():
	options = Options()
	prefs = {
		"profile.default_content_setting_values.plugins": 1,
		"profile.content_settings.plugin_whitelist.adobe-flash-player": 1,
		"profile.content_settings.exceptions.plugins.*,*.per_resource.adobe-flash-player": 1,
		"PluginsAllowedForUrls": "https://kof.bizsys.pearson.com/markview/MVT_Web_Inquiry.ShowInquiry"
	}
	options.add_experimental_option('prefs',prefs)
	# options.add_argument('--start-maximized')
	return wd.Chrome('./chromedriver',options=options)
	# return wd.Ie(executable_path='IEDriverServer')
	# bina = FirefoxBinary('C:\\Program Files\\Mozilla Firefox\\firefox.exe')
	# return wd.Firefox(firefox_binary=bina)

D = Drive()




def StaRt():
	D.get('https://kof.bizsys.pearson.com/markview/MVT_Web_Inquiry.Home?InquiryType=APINVOICE&HelpID=Web%20Inquiry%20-%20AP%20Invoices')
	# D.get('https://pearson.coupahost.com/invoices')
	# D.get('http://localhost:7000/')
	# time.sleep(5)
	timeout = 10
	try:
		wdw(D, timeout).until(ec.visibility_of_element_located((By.XPATH, '//*[@id="user-name-txt"]')))
		print('Complete MFA')

	except te:
		print('Time Up, closing application, please re-run')
		D.quit()

	timeout = 61
	try:
		wdw(D, timeout).until(ec.visibility_of_element_located((By.XPATH, '//*[@id="82"]')))
		# wdw(D, timeout).until(ec.visibility_of_element_located((By.XPATH, '//*[@id="pageHeader"]')))

	except te:
		print('You Did not Enter the MFA Code')
		D.quit()
StaRt()


# DummyInput = ['7027880','HEINEMANN',' ','01-30-2019','USD',' ','30','71010002357']

def HA(h): # Invoice Number
	temp = []
	for a,b in enumerate(h):
		temp.append(h[a])
	return temp

Ha = HA(ValidPO)

# pprint(Ha)

newData = []
newDataVen = {}
OneData = []
MarkOutData = []
invnu = []
def getData(Data,ClenData,venderData,VenDataAdd,Check):
	Fetch = []
	y = 16
	for dat,Dat in enumerate(Data):
		td = Dat.find_elements_by_tag_name('td')
		if len(td) == 16:
			if Check == 'Mark':
				if dat == 0:
					continue
			for x,xx in enumerate(td):
				Fetch.append(xx.get_attribute('innerText'))
	CD = [Fetch[x:x+16] for x in range(0, len(Fetch),16)]

	VenDataAdd[venderData[0]] = [[venderData[0],venderData[6]],CD]

	for a in CD:
		ClenData.append(a)
		# print(a)
		# print(' ')
		# print(' ')
		# print(' ')
		# print(' ')
		# print(' ')


def FetchData(H):
	data = D.find_element_by_xpath('/html/body/table/tbody/tr[3]/td/table[3]/tbody/tr[1]/td/table[2]/tbody')
	dat = data.find_elements_by_tag_name('tr')

	getData(dat,newData,H,newDataVen,'Mark')

# FetchData(DummyInput)


def MarSta(getDa,data):
	for md,MD in enumerate(data):
		VendorInvoice = MD
		for dda,fdda in enumerate(getDa):
			# print('MD'*100)
			# print(fdda)
			# print('MD'*100)
			# print('')
			# print('')
			# print('-'*100)
			# print('')
			# print('')
			# print('VD'*100)
			# print(data)
			# print('VD'*100)
			if (fdda[13] == 'Completed' and fdda[1] != '0.00' and fdda[1][:1] != '-' and fdda[12] != 'VOIDED') or (fdda[13] == 'Archive' and fdda[12] != 'VOIDED') or (fdda[13] == 'Archive' and fdda[12] != '\xa0') or (fdda[13] == 'Comment Request' and fdda[8] != '\xa0' and fdda[12] != 'VOIDED') or (fdda[12] == 'NEGOTIABLE' and fdda[8] != '\xa0'):	
				SelectDataM = [fdda[5],fdda[6],fdda[7],fdda[1]]
				SelectDataM.append('PAI')
				SelectDataM.append('PAI')
				SelectDataMEx = [fdda[5],fdda[6],fdda[7],fdda[1]]
				SelectDataMEx.append('Paid via ACH# {} on {}'.format(fdda[8],fdda[9]))
				SendDataM = {}
				# SendDataM[VendorInvoice] = data[VendorInvoice][0],SelectDataM,SelectDataMEx
				SendDataM[fdda[5]] = data[VendorInvoice][0],SelectDataM,SelectDataMEx
				OneData.append(SendDataM)			
				MarkOutData.append(fdda)

			elif fdda[13] == 'Completed' and (fdda[1][:1] == '-'):
				SelectDataM = [fdda[5],fdda[6],fdda[7],data[VendorInvoice][0][1]]
				SelectDataM.append('OTH')
				SelectDataM.append('OTH')
				SelectDataMEx = [fdda[5],fdda[6],fdda[7],data[VendorInvoice][0][1]]
				SelectDataMEx.append('Others')
				SendDataM = {}
				# SendDataM[VendorInvoice] = data[VendorInvoice][0],SelectDataM,SelectDataMEx
				SendDataM[fdda[5]] = data[VendorInvoice][0],SelectDataM,SelectDataMEx
				OneData.append(SendDataM)
				MarkOutData.append(fdda)

			elif fdda[13] == 'Cancelled' or (fdda[13] == 'Completed' and fdda[1] == '0.00' and fdda[1][:1] != '-'):
				SelectDataM = [fdda[5],fdda[6],fdda[7],data[VendorInvoice][0][1]]
				SelectDataM.append('OTH')
				SelectDataM.append('OTH')
				SelectDataMEx = [fdda[5],fdda[6],fdda[7],data[VendorInvoice][0][1]]
				SelectDataMEx.append('Others')				
				SendDataM = {}
				# SendDataM[VendorInvoice] = data[VendorInvoice][0],SelectDataM,SelectDataMEx
				SendDataM[fdda[5]] = data[VendorInvoice][0],SelectDataM,SelectDataMEx
				OneData.append(SendDataM)
				MarkOutData.append(fdda)

			elif (fdda[13] == 'Pending Payment' and fdda[1] != '0.00') or fdda[12] == 'VOIDED':
				SelectDataM = [fdda[5],fdda[6],fdda[7],fdda[1]]
				SelectDataM.append('RFP')
				SelectDataM.append('RFP')
				SelectDataMEx = [fdda[5],fdda[6],fdda[7],fdda[1]]
				SelectDataMEx.append('Will be paid as per due date')
				SendDataM = {}
				# SendDataM[VendorInvoice] = data[VendorInvoice][0],SelectDataM,SelectDataMEx
				SendDataM[fdda[5]] = data[VendorInvoice][0],SelectDataM,SelectDataMEx
				OneData.append(SendDataM)
				MarkOutData.append(fdda)

			elif (fdda[13] == 'Comment Request' and fdda[8] == '\xa0') or fdda[13] == 'Approval and Coding' or fdda[13] == 'AP PO Invoice Process Resolution' or fdda[13] == 'AP Resolution':
				SelectDataM = [fdda[5],fdda[6],fdda[7],fdda[1]]
				SelectDataM.append('PWB')
				SelectDataM.append('PWB')
				SelectDataMEx = [fdda[5],fdda[6],fdda[7],fdda[1]]
				SelectDataMEx.append('Pending with Pearson business')
				SendDataM = {}
				# SendDataM[VendorInvoice] = data[VendorInvoice][0],SelectDataM,SelectDataMEx
				SendDataM[fdda[5]] = data[VendorInvoice][0],SelectDataM,SelectDataMEx
				OneData.append(SendDataM)
				MarkOutData.append(fdda)

			elif fdda[13] == 'Receiving Resolution':
				SelectDataM = [fdda[5],fdda[6],fdda[7],fdda[1]]
				SelectDataM.append('PWB')
				SelectDataM.append('PWB')
				SelectDataMEx = [fdda[5],fdda[6],fdda[7],fdda[1]]
				SelectDataMEx.append('Pending with Pearson business')
				SendDataM = {}
				# SendDataM[VendorInvoice] = data[VendorInvoice][0],SelectDataM,SelectDataMEx
				SendDataM[fdda[5]] = data[VendorInvoice][0],SelectDataM,SelectDataMEx
				OneData.append(SendDataM)
				MarkOutData.append(fdda)

			elif fdda[13] == 'Verify AP Process Complete':
				SelectDataM = [fdda[5],fdda[6],fdda[7],fdda[1]]
				SelectDataM.append('PWB')
				SelectDataM.append('PWB')
				SelectDataMEx = [fdda[5],fdda[6],fdda[7],fdda[1]]
				SelectDataMEx.append('Pending with Pearson business')
				SendDataM = {}
				# SendDataM[VendorInvoice] = data[VendorInvoice][0],SelectDataM,SelectDataMEx
				SendDataM[fdda[5]] = data[VendorInvoice][0],SelectDataM,SelectDataMEx
				OneData.append(SendDataM)
				MarkOutData.append(fdda)

			elif fdda[13] == 'Senior Financial Audit':
				SelectDataM = [fdda[5],fdda[6],fdda[7],fdda[1]]
				SelectDataM.append('PWB')
				SelectDataM.append('PWB')
				SelectDataMEx = [fdda[5],fdda[6],fdda[7],fdda[1]]
				SelectDataMEx.append('Pending with Pearson business')
				SendDataM = {}
				# SendDataM[VendorInvoice] = data[VendorInvoice][0],SelectDataM,SelectDataMEx
				SendDataM[fdda[5]] = data[VendorInvoice][0],SelectDataM,SelectDataMEx
				OneData.append(SendDataM)
				MarkOutData.append(fdda)


			elif fdda[4] == 'Approved' and fdda[8] == 'No':
				SelectDataM = [fdda[0],fdda[2],fdda[15],fdda[3].split(' ')[0]]
				SelectDataM.append('EIC')
				SelectDataM.append('EIC')
				SelectDataMEx = [fdda[0],fdda[2],fdda[15],fdda[3].split(' ')[0]]
				SelectDataMEx.append('Pending in Coupa')
				SendDataM = {}
				# SendDataM[VendorInvoice] = data[VendorInvoice][0],SelectDataM,SelectDataMEx
				SendDataM[fdda[0]] = data[VendorInvoice][0],SelectDataM,SelectDataMEx
				OneData.append(SendDataM)
				MarkOutData.append(fdda)


			elif fdda[4] == 'Approved' and fdda[8] == 'Yes' and fdda[7] != 'None':
				SelectDataM = [fdda[0],fdda[2],fdda[15],fdda[3].split(' ')[0]]
				SelectDataM.append('EIC')
				SelectDataM.append('EIC')
				SelectDataMEx = [fdda[0],fdda[2],fdda[15],fdda[3].split(' ')[0]]
				SelectDataMEx.append('Pending in Coupa')
				SendDataM = {}
				# SendDataM[VendorInvoice] = data[VendorInvoice][0],SelectDataM,SelectDataMEx
				SendDataM[fdda[0]] = data[VendorInvoice][0],SelectDataM,SelectDataMEx
				OneData.append(SendDataM)
				MarkOutData.append(fdda)


			elif fdda[4] == 'Approved' and fdda[8] == 'Yes' and fdda[7] == 'None':
				SelectDataM = [fdda[0],fdda[2],fdda[15],fdda[3].split(' ')[0]]
				SelectDataM.append('EIC')
				SelectDataM.append('EIC')
				SelectDataMEx = [fdda[0],fdda[2],fdda[15],fdda[3].split(' ')[0]]
				SelectDataMEx.append('Pending in Coupa')
				SendDataM = {}
				# SendDataM[VendorInvoice] = data[VendorInvoice][0],SelectDataM,SelectDataMEx
				SendDataM[fdda[0]] = data[VendorInvoice][0],SelectDataM,SelectDataMEx
				OneData.append(SendDataM)
				MarkOutData.append(fdda)


			elif fdda[4] == 'Pending Receipt':
				SelectDataM = [fdda[0],fdda[2],fdda[15],fdda[3].split(' ')[0]]
				SelectDataM.append('PWB')
				SelectDataM.append('PWB')
				SelectDataMEx = [fdda[0],fdda[2],fdda[15],fdda[3].split(' ')[0]]
				SelectDataMEx.append('Pending with Pearson business')
				SendDataM = {}
				# SendDataM[VendorInvoice] = data[VendorInvoice][0],SelectDataM,SelectDataMEx
				SendDataM[fdda[0]] = data[VendorInvoice][0],SelectDataM,SelectDataMEx
				OneData.append(SendDataM)
				MarkOutData.append(fdda)


			elif fdda[4] == 'Draft':
				SelectDataM = [fdda[0],fdda[2],fdda[15],fdda[3].split(' ')[0]]
				SelectDataM.append('EIC')
				SelectDataM.append('EIC')
				SelectDataMEx = [fdda[0],fdda[2],fdda[15],fdda[3].split(' ')[0]]
				SelectDataMEx.append('Pending in Coupa')
				SendDataM = {}
				# SendDataM[VendorInvoice] = data[VendorInvoice][0],SelectDataM,SelectDataMEx
				SendDataM[fdda[0]] = data[VendorInvoice][0],SelectDataM,SelectDataMEx
				OneData.append(SendDataM)
				MarkOutData.append(fdda)
			
			elif fdda[4] == 'Disputed':
				SelectDataM = [fdda[0],fdda[2],fdda[15],fdda[3].split(' ')[0]]
				SelectDataM.append('EIC')
				SelectDataM.append('EIC')
				SelectDataMEx = [fdda[0],fdda[2],fdda[15],fdda[3].split(' ')[0]]
				SelectDataMEx.append('Pending in Coupa')
				SendDataM = {}
				# SendDataM[VendorInvoice] = data[VendorInvoice][0],SelectDataM,SelectDataMEx
				SendDataM[fdda[0]] = data[VendorInvoice][0],SelectDataM,SelectDataMEx
				OneData.append(SendDataM)
				MarkOutData.append(fdda)
			
			elif fdda[5] == 'Coupa Supplier Portal' and fdda[4] != 'Disputed' and fdda[4] != 'Draft' and fdda[4] != 'Pending Receipt' and fdda[4] != 'Approved':
				SelectDataM = [fdda[0],fdda[2],fdda[15],fdda[3].split(' ')[0]]
				SelectDataM.append('OTH')
				SelectDataM.append('OTH')
				SelectDataMEx = [fdda[0],fdda[2],fdda[15],fdda[3].split(' ')[0]]
				SelectDataMEx.append('Others')
				SendDataM = {}
				# SendDataM[VendorInvoice] = data[VendorInvoice][0],SelectDataM,SelectDataMEx
				SendDataM[fdda[0]] = data[VendorInvoice][0],SelectDataM,SelectDataMEx
				OneData.append(SendDataM)
				MarkOutData.append(fdda)

			else:
				SelectDataM = [fdda[5],fdda[6],fdda[7],data[VendorInvoice][0][1]]
				SelectDataM.append('OTH')
				SelectDataM.append('OTH')
				SelectDataMEx = [fdda[5],fdda[6],fdda[7],data[VendorInvoice][0][1]]
				SelectDataMEx.append('Others')				
				SendDataM = {}
				# SendDataM[VendorInvoice] = data[VendorInvoice][0],SelectDataM,SelectDataMEx
				SendDataM[fdda[5]] = data[VendorInvoice][0],SelectDataM,SelectDataMEx
				OneData.append(SendDataM)
				MarkOutData.append(fdda)




def Single_Invoice_Date(data):
	# print('Sing Invoice Found')
	# print(data)
	for md,MD in enumerate(data):
		VendorInvoice = MD
		for a,aa in enumerate(data[VendorInvoice][1]):
			MarSta(data[VendorInvoice][1],data)
	
	newData.clear()	
	newDataVen.clear()

def Multi_Invoice_Data(data):
	# print('Multiple Invoice Found')
	CancelledI = False
	VoidedI = False
	CancelledO = False
	VoidedO = False	
	VoidedandPaid = False
	CancelledandReprocessed = False
	Reprocessed = []
	tempLD = {}
	tempLN = {}
	temp = []
	temP = []	
	for md,MD in enumerate(data):
		VendorInvoice = MD
		# print(VendorInvoice)
		for a,aa in enumerate(data[VendorInvoice][1]):
			if (aa[13] == 'Cancelled') or (aa[13] == 'Returned Invoices' and aa[1] == '0.00') or (aa[13] == '\xa0' and aa[1] == '0.00'):
				CancelledI = True
				# print('CancelledI True')
				# print(aa)
				# print('CancelledI True')
				# print('')
				# print('')
				# print('')				
			else:
				CancelledandReprocessed = True
				Reprocessed.append(aa)
				# print('CancelledandReprocessed True	')
				# print(aa)
				# print('CancelledandReprocessed True	')
				# print('')
				# print('')
				# print('')
		
		if CancelledandReprocessed:
			for b,bb in enumerate(Reprocessed):

				if bb[12] == 'VOIDED':
					VoidedI = True
				else:
					temp.append(bb)
				
				if VoidedI:
					for cv,ccv in enumerate(temp):
						# print(ccv)
						if (ccv[13] == 'Completed' and ccv[1] != '0.00' and ccv[1][:1] != '-' and ccv[12] != 'VOIDED') or (ccv[13] == 'Archive' and ccv[12] != 'VOIDED') or (ccv[13] == 'Archive' and ccv[12] != '\xa0') or (ccv[13] == 'Comment Request' and ccv[8] != '\xa0' and ccv[12] != 'VOIDED'):
							VoidedandPaid = True
							VoidedI = False
							VoidedO = True
							break
						else:
							VoidedandPaid = False
							VoidedI = False
							VoidedO = True
						if VoidedI == False and VoidedandPaid == True:
							if (ccv[13] == 'Completed' and ccv[1] != '0.00' and ccv[1][:1] != '-' and ccv[12] != 'VOIDED') or (ccv[13] == 'Archive' and ccv[12] != 'VOIDED') or (ccv[13] == 'Archive' and ccv[12] != '\xa0') or (ccv[13] == 'Comment Request' and ccv[8] != '\xa0' and ccv[12] != 'VOIDED'):
								SelectData = [ccv[5],ccv[6],ccv[7],ccv[1]]
								SelectData.append('PAI')
								SelectData.append('PAI')
								SelectDataEx = [ccv[5],ccv[6],ccv[7],ccv[1]]
								SelectDataEx.append('Paid via ACH# {} on {}'.format(ccv[8],ccv[9]))
								SendData = {}
								SendData[ccv[5]] = data[VendorInvoice][0],SelectData,SelectDataEx
								OneData.append(SendData)
								MarkOutData.append(ccv)
								break
							elif VoidedI == False and VoidedandPaid == False:
								SelectData = [ccv[5],ccv[6],ccv[7],ccv[1]]
								SelectData.append('RFP')
								SelectData.append('RFP')
								SelectDataEx = [ccv[5],ccv[6],ccv[7],ccv[1]]
								SelectDataEx.append('Will be paid as per due date')
								SendData = {}
								SendData[ccv[5]] = data[VendorInvoice][0],SelectData,SelectDataEx
								OneData.append(SendData)
								MarkOutData.append(ccv)
								break					


			if len(temp) > 1:
				for Q,QQ in enumerate(temp):
					X = Q
					for R,RR in enumerate(temp):
						if X == R:
							continue
						if temp[Q][1] == temp[R][1] and len(temp[Q][5]) > len(temp[R][5]):
							tempLD[QQ[1]] = QQ
						else:
							tempLN[QQ[1]] = QQ

				if len(tempLD) != 0:
					for tl,lt in enumerate(tempLD.keys()):
						temP.append(tempLD[lt])

				if len(tempLN) != 0:
						
					for tl,lt in enumerate(tempLN.keys()):
						temP.append(tempLN[lt])
				
				NOZERO = False
				# if len(temP) > 1:
				for ta,tta in enumerate(temP):
					if temP[ta][1] == '0.00':
						NOZERO = True


				if NOZERO:
					for t,tt in enumerate(temP):
						if temP[t][1] == '0.00':
							continue
						else:
							# print('temP')
							# print(temP)
							# print(' ')
							# print(' ')
							# print(' ')
							# print(' ')
							MarSta([tt],data)
				else:
					SelectData = [data[VendorInvoice][1][0][5],data[VendorInvoice][1][0][6],data[VendorInvoice][1][0][7],data[VendorInvoice][0][1]]
					SelectData.append('OTH')
					SelectData.append('OTH')
					SelectDataEx = [data[VendorInvoice][1][0][5],data[VendorInvoice][1][0][6],data[VendorInvoice][1][0][7],data[VendorInvoice][0][1]]
					SelectDataEx.append('Others')			
					SendData = {}
					SendData[data[VendorInvoice][1][0][5]] = data[VendorInvoice][0],SelectData,SelectDataEx
					OneData.append(SendData)
					MarkOutData.append(data[VendorInvoice][1])
			elif len(temp) == 1:
				# print(temp)
				MarSta(temp,data)

		else:
			# print('CancelledandReprocessed False')
			# print(aa)
			# print('CancelledandReprocessed False')
			# print('')
			# print('')
			# print('')

			SelectData = [data[VendorInvoice][1][0][5],data[VendorInvoice][1][0][6],data[VendorInvoice][1][0][7],data[VendorInvoice][0][1]]
			SelectData.append('OTH')
			SelectData.append('OTH')
			SelectDataEx = [data[VendorInvoice][1][0][5],data[VendorInvoice][1][0][6],data[VendorInvoice][1][0][7],data[VendorInvoice][0][1]]
			SelectDataEx.append('Others')			
			SendData = {}
			SendData[data[VendorInvoice][1][0][5]] = data[VendorInvoice][0],SelectData,SelectDataEx
			OneData.append(SendData)
			MarkOutData.append(data[VendorInvoice][1])

	newData.clear()
	newDataVen.clear()

def sivin():
	print('Searching invoices with invoice number only.')
	for P,H in enumerate(Ha):
		D.find_element_by_xpath('//*[@id="82"]').send_keys(H[0]+'%')
		D.find_element_by_xpath('/html/body/table/tbody/tr[3]/td/table[2]/tbody/tr/td/table/tbody/tr/td[2]/table/tbody/tr[5]/td[1]/table/tbody/tr/td[1]/div/a').send_keys(Keys.ENTER)
		# time.sleep(1)
		if D.find_element_by_xpath('/html/body/table/tbody/tr[3]/td/table[3]/tbody/tr[2]/td').get_attribute('innerText') == 'Query returned no records.':
			invnu.append(H)
			D.find_element_by_xpath('//*[@id="82"]').clear()
		else:
			data = D.find_element_by_xpath('/html/body/table/tbody/tr[3]/td/table[3]/tbody/tr[1]/td/table[2]/tbody')
			dat = data.find_elements_by_tag_name('tr')
			getData(dat,newData,H,newDataVen,'Mark')

			if len(newData) > 1:
				Multi_Invoice_Data(newDataVen)
			else:
				Single_Invoice_Date(newDataVen)
			
			D.find_element_by_xpath('//*[@id="82"]').clear()
		print(str(P+1)+'. '+H[0])




def sivinav():
	print('Searching invoices with vendor name and number')
	for P,H in enumerate(Ha):
		D.find_element_by_xpath('//*[@id="82"]').send_keys(H[0]+'%')
		D.find_element_by_xpath('//*[@id="94"]').send_keys(H[1])
		D.find_element_by_xpath('/html/body/table/tbody/tr[3]/td/table[2]/tbody/tr/td/table/tbody/tr/td[2]/table/tbody/tr[5]/td[1]/table/tbody/tr/td[1]/div/a').send_keys(Keys.ENTER)
		time.sleep(0.7)
		if D.find_element_by_xpath('/html/body/table/tbody/tr[3]/td/table[3]/tbody/tr[2]/td').get_attribute('innerText') == 'Query returned no records.':
			invnu.append(H)
			D.find_element_by_xpath('//*[@id="82"]').clear()
			D.find_element_by_xpath('//*[@id="94"]').clear()
		else:
			data = D.find_element_by_xpath('/html/body/table/tbody/tr[3]/td/table[3]/tbody/tr[1]/td/table[2]/tbody')
			dat = data.find_elements_by_tag_name('tr')
			getData(dat,newData,H,newDataVen,'Mark')

			if len(newData) > 1:
				Multi_Invoice_Data(newDataVen)
			else:
				Single_Invoice_Date(newDataVen)
			
			D.find_element_by_xpath('//*[@id="82"]').clear()
			D.find_element_by_xpath('//*[@id="94"]').clear()
		print(str(P+1)+'. '+H[0])

conNcon(IVN)

CoupaInv = []
MarkInv = []
def Cou():
	for i,ii in enumerate(invnu):
		if ii[7][:3] == 'COU':
			CoupaInv.append(ii)
		else:
			MarkInv.append(ii)
Cou()


# def Cou():
# 	for i,ii in enumerate(ValidPO):
# 		if ii[7][:3] == 'COU':
# 			CoupaInv.append(ii)
# 		else:
# 			MarkInv.append(ii)
# Cou()

COUPRIC = []
def Cfci():
		D.get('https://pearson.coupahost.com/invoices')
		timeout = 10
		try:
			wdw(D, timeout).until(ec.visibility_of_element_located((By.XPATH, '//*[@id="pageHeader"]')))
			D.find_element_by_xpath('//*[@id="invoice_header_filter"]/option[2]').click()

			for co,coo in enumerate(CoupaInv):
				D.find_element_by_xpath('//*[@id="sf_invoice_header"]').send_keys(coo[0])
				D.find_element_by_xpath('//*[@id="invoice_header_data_table_form_search"]/div[1]/table/tbody/tr/td[2]/table/tbody/tr/td[7]/div/a').send_keys(Keys.ENTER)
				time.sleep(7)
				if D.find_element_by_xpath('//*[@id="invoice_header_tbody"]/tr/td').get_attribute('innerText') == 'Nothing matching your search was found.':
					COUPRIC.append(coo)
					D.find_element_by_xpath('//*[@id="sf_invoice_header"]').clear()
				else:
					data = D.find_element_by_xpath('//*[@id="invoice_header_tbody"]')
					dat = data.find_elements_by_tag_name('tr')

					getData(dat,newData,coo,newDataVen,'Coup')

					if len(newData) > 1:
						Multi_Invoice_Data(newDataVen)
					else:
						Single_Invoice_Date(newDataVen)

					D.find_element_by_xpath('//*[@id="sf_invoice_header"]').clear()
				print(str(co+1)+'. '+coo[0])
		except te:
			print('There is problem in COUPA portal, please check and re-run')
			D.quit()					

if len(CoupaInv) !=0:
	conNcon(IZI)

def reqInv(R):
	for a,aa in enumerate(R):
		SelectData = [aa[0],aa[3],aa[4],aa[6]]
		SelectData.append('RIC')
		SelectData.append('RIC')
		SelectDataMEx = [aa[0],aa[3],aa[4],aa[6]]
		SelectDataMEx.append('Required invoice copy')
		SendData = {}
		SendData[aa[0]] = [aa[0],aa[6]],SelectData,SelectDataMEx
		OneData.append(SendData)
		MarkOutData.append(aa)

if len(MarkInv) !=0:
	reqInv(MarkInv)

if len(COUPRIC) !=0:
	reqInv(COUPRIC)

def reqInv(R):
	for a,aa in enumerate(R):
		SelectData = [aa[0],aa[3],aa[4],aa[6]]
		SelectData.append('NUS')
		SelectData.append('NUS')
		SelectDataMEx = [aa[0],aa[3],aa[4],aa[6]]
		SelectDataMEx.append('PO does not pertain to US')		
		SendData = {}
		SendData[aa[0]] = [aa[0],aa[6]],SelectData,SelectDataMEx
		OneData.append(SendData)
		MarkOutData.append(aa)

if len(InValidPO) !=0:
	reqInv(InValidPO)


# pprint(OneData)


def ConvertFloat(Data):
	# print(Data)
	# print(len(Data),'             ', type(Data))
	# print(Data.index(1))
	# print(' ')
	# print(' ')
	# print(' ')
	# print(' ')
	# print(' ')
	for ll in Data:
		# print(ll)
		# print(type(ll))
		for x in ll.items():
			if type(x[1][0][1]) == str:
				x[1][0][1] = round(float(x[1][0][1].replace(',','')),2)
			if type(x[1][1][3]) == str:
				x[1][1][3] = round(float(x[1][1][3].replace(',','')),2)
			if type(x[1][2][3]) == str:
				x[1][2][3] = round(float(x[1][2][3].replace(',','')),2)
			# print(x)
		# 	print(x[1][0])
				# x[1][0].append(400.00)
		# 	print(x)
		# 	break
			
			# print(x[1][0][1])
			# print(type(x[1][0][1]))
			# print(ll, '          ', type(ll[x][0][1]))
			# x[1][0][1] = round(float(x[1][0][1].replace(',','')),2)
			# ll[x][1][3] = round(float(ll[x][1][3].replace(',','')),2)
			# ll[x][2][3] = round(float(ll[x][2][3].replace(',','')),2)
			# print(ll, '          ', type(ll[x][0][1]))
	# print(Data)
ConvertFloat(OneData)

# pprint(OneData)



def CheckAmount(AMT):
	for l,ll in enumerate(AMT):
		for x in ll.keys():
			if ll[x][0][1] != ll[x][1][3]:
				ll[x][1][5] = 'DAM'
		# print(ll)
CheckAmount(OneData)

# pprint(OneData)


# print(len(CoupaInv))
# print(CoupaInv)

# if len(newData) > 1:
# 	Multi_Invoice_Data(newDataVen)
# 	ConvertFloat(OneData)
# 	CheckAmount(OneData)
# else:
# 	Single_Invoice_Date(newDataVen)
# 	ConvertFloat(OneData)
# 	CheckAmount(OneData)

# print(OneData)

def VenAmtADD(e):
	Sum = 0.00
	for x in e:
		Sum = x[1] + Sum
	return Sum

def ADD(e):
	Sum = 0.00
	for x in e:
		Sum = x[4] + Sum
	return Sum

def ADDDAMA1(e):
	Sum = 0.00
	for x in e:
		Sum = x[3] + Sum
	return Sum

def ADDDAMA2(e):
	Sum = 0.00
	for x in e:
		Sum = x[1] + Sum
	return Sum



wk = XL.Workbook()
WB = Workbook()



DEFHID = ['Sl No.','Invoice#','Invoice Date','Invoice Currency','Invoice Amount','Status']
DEFHIDDA = ['Sl No.','Invoice#','Invoice Date','Invoice Currency','Invoice Amount','Actual Amount','Status']

Tven = 0.00
Tpai = 0.00
Tpwb = 0.00
Tric = 0.00
Trfp = 0.00
Teic = 0.00
Tnus = 0.00
Tdam = 0.00
Toth = 0.00


def WriteTabs(wrt):

	sh = WB.create_sheet('Paid',1)
	sh = WB.create_sheet('Ready for Payment',2)
	sh = WB.create_sheet('Pending with Pearson',3)
	sh = WB.create_sheet('Required Invoice Copy',4)
	sh = WB.create_sheet('Pending in COUPA',5)
	sh = WB.create_sheet('Not US Invoice',6)
	sh = WB.create_sheet('Pending with Pearson AP',7)
	sh = WB.create_sheet('Cancelled',8)
	sh = WB.create_sheet('Difference Amount Recorded',9)
	sh = WB.create_sheet('Others',10)

	tempData = {}
	tempDataT = []
	tempDataN = []

	for X,XX in enumerate(wrt):
		for x in XX.keys():
			if XX[x][1][5] == 'PAI':
				tempData[XX[x][1][0]] = XX[x][2]
				tempDataT.append(XX[x][2])
				sk = WB.active = WB['Paid']
				for x,xx in enumerate(DEFHID):
					sk.cell(column=(x+1),row=(1),value=DEFHID[x])
	for pa,pA in enumerate(tempData.keys()):
		tempData[pA].insert(0,pa+1)
		for pp,pP in enumerate(tempData[pA]):
			sk.cell(column=(pp+1),row=(pa+2),value=pP)
			sk.cell(column=(5),row=(pa+2),value=tempData[pA][4])
	for X,XX in enumerate(wrt):		
		for x in XX.keys():
			if XX[x][1][5] == 'PAI':
				sk.merge_cells('C{0}:D{1}'.format(len(tempData)+2,len(tempData)+2))
				sk.cell(column=(3),row=(len(tempData)+2),value='Total')
				sk.cell(column=(5),row=(len(tempData)+2),value=ADD(tempDataT))
	Tpai = ADD(tempDataT)
	tempData.clear()
	tempDataT.clear()


	for X,XX in enumerate(wrt):
		for x in XX.keys():
			if XX[x][1][5] == 'RFP':
				tempData[XX[x][1][0]] = XX[x][2]
				tempDataT.append(XX[x][2])
				sk = WB.active = WB['Ready for Payment']
				for x,xx in enumerate(DEFHID):
					sk.cell(column=(x+1),row=(1),value=DEFHID[x])
	for pa,pA in enumerate(tempData.keys()):
		tempData[pA].insert(0,pa+1)
		for pp,pP in enumerate(tempData[pA]):
			sk.cell(column=(pp+1),row=(pa+2),value=pP)
			sk.cell(column=(5),row=(pa+2),value=tempData[pA][4])
	for X,XX in enumerate(wrt):			
		for x in XX.keys():
			if XX[x][1][5] == 'RFP':
				sk.merge_cells('C{0}:D{1}'.format(len(tempData)+2,len(tempData)+2))		
				sk.cell(column=(3),row=(len(tempData)+2),value='Total')
				sk.cell(column=(5),row=(len(tempData)+2),value=ADD(tempDataT))
	Trfp = ADD(tempDataT)
	tempData.clear()
	tempDataT.clear()

	for X,XX in enumerate(wrt):
		for x in XX.keys():
			if XX[x][1][5] == 'PWB':
				tempData[XX[x][1][0]] = XX[x][2]
				tempDataT.append(XX[x][2])
				sk = WB.active = WB['Pending with Pearson']
				for x,xx in enumerate(DEFHID):
					sk.cell(column=(x+1),row=(1),value=DEFHID[x])
	for pa,pA in enumerate(tempData.keys()):
		tempData[pA].insert(0,pa+1)
		for pp,pP in enumerate(tempData[pA]):
			sk.cell(column=(pp+1),row=(pa+2),value=pP)
			sk.cell(column=(5),row=(pa+2),value=tempData[pA][4])
	for X,XX in enumerate(wrt):			
		for x in XX.keys():
			if XX[x][1][5] == 'PWB':
				sk.merge_cells('C{0}:D{1}'.format(len(tempData)+2,len(tempData)+2))		
				sk.cell(column=(3),row=(len(tempData)+2),value='Total')
				sk.cell(column=(5),row=(len(tempData)+2),value=ADD(tempDataT))
	Tpwb = ADD(tempDataT)
	tempData.clear()
	tempDataT.clear()		


	for X,XX in enumerate(wrt):
		for x in XX.keys():
			if XX[x][1][5] == 'RIC':
				tempData[XX[x][1][0]] = XX[x][2]
				tempDataT.append(XX[x][2])
				sk = WB.active = WB['Required Invoice Copy']
				for x,xx in enumerate(DEFHID):
					sk.cell(column=(x+1),row=(1),value=DEFHID[x])
	for pa,pA in enumerate(tempData.keys()):
		tempData[pA].insert(0,pa+1)
		for pp,pP in enumerate(tempData[pA]):
			sk.cell(column=(pp+1),row=(pa+2),value=pP)
			sk.cell(column=(5),row=(pa+2),value=tempData[pA][4])
	for X,XX in enumerate(wrt):			
		for x in XX.keys():
			if XX[x][1][5] == 'RIC':
				sk.merge_cells('C{0}:D{1}'.format(len(tempData)+2,len(tempData)+2))		
				sk.cell(column=(3),row=(len(tempData)+2),value='Total')
				sk.cell(column=(5),row=(len(tempData)+2),value=ADD(tempDataT))
	Tric = ADD(tempDataT)
	tempData.clear()
	tempDataT.clear()


	for X,XX in enumerate(wrt):
		for x in XX.keys():
			if XX[x][1][5] == 'EIC':
				tempData[XX[x][1][0]] = XX[x][2]
				tempDataT.append(XX[x][2])
				sk = WB.active = WB['Pending in COUPA']
				for x,xx in enumerate(DEFHID):
					sk.cell(column=(x+1),row=(1),value=DEFHID[x])
	for pa,pA in enumerate(tempData.keys()):
		tempData[pA].insert(0,pa+1)
		for pp,pP in enumerate(tempData[pA]):
			sk.cell(column=(pp+1),row=(pa+2),value=pP)
			sk.cell(column=(5),row=(pa+2),value=tempData[pA][4])
	for X,XX in enumerate(wrt):			
		for x in XX.keys():
			if XX[x][1][5] == 'EIC':
				sk.merge_cells('C{0}:D{1}'.format(len(tempData)+2,len(tempData)+2))		
				sk.cell(column=(3),row=(len(tempData)+2),value='Total')
				sk.cell(column=(5),row=(len(tempData)+2),value=ADD(tempDataT))
	Teic = ADD(tempDataT)
	tempData.clear()
	tempDataT.clear()

	for X,XX in enumerate(wrt):
		for x in XX.keys():
			if XX[x][1][5] == 'NUS':
				tempData[XX[x][1][0]] = XX[x][2]
				tempDataT.append(XX[x][2])
				sk = WB.active = WB['Not US Invoice']
				for x,xx in enumerate(DEFHID):
					sk.cell(column=(x+1),row=(1),value=DEFHID[x])
	for pa,pA in enumerate(tempData.keys()):
		tempData[pA].insert(0,pa+1)
		for pp,pP in enumerate(tempData[pA]):
			sk.cell(column=(pp+1),row=(pa+2),value=pP)
			sk.cell(column=(5),row=(pa+2),value=tempData[pA][4])
	for X,XX in enumerate(wrt):			
		for x in XX.keys():
			if XX[x][1][5] == 'NUS':
				sk.merge_cells('C{0}:D{1}'.format(len(tempData)+2,len(tempData)+2))		
				sk.cell(column=(3),row=(len(tempData)+2),value='Total')
				sk.cell(column=(5),row=(len(tempData)+2),value=ADD(tempDataT))
	Tnus = ADD(tempDataT)
	tempData.clear()
	tempDataT.clear()		

	for X,XX in enumerate(wrt):
		for x in XX.keys():
			if XX[x][1][5] == 'DAM':
				DAMWrite = [XX[x][1][0],XX[x][1][1],XX[x][1][2],XX[x][1][3],XX[x][0][1],'Diffrenet Amount']
				# print(DAMWrite)
				tempData[XX[x][1][0]] = DAMWrite
				# tempDataT.append(DAMWrite)
				tempDataT.append(XX[x][2])
				tempDataN.append(XX[x][0])
				sk = WB.active = WB['Difference Amount Recorded']
				for x,xx in enumerate(DEFHIDDA):
					sk.cell(column=(x+1),row=(1),value=DEFHIDDA[x])
	for pa,pA in enumerate(tempData.keys()):
		# print(tempData[pA])
		tempData[pA].insert(0,pa+1)
		for pp,pP in enumerate(tempData[pA]):
			sk.cell(column=(pp+1),row=(pa+2),value=pP)
			sk.cell(column=(5),row=(pa+2),value=tempData[pA][4])
			sk.cell(column=(6),row=(pa+2),value=tempData[pA][5])
	for X,XX in enumerate(wrt):			
		for x in XX.keys():
			if XX[x][1][5] == 'DAM':
				sk.merge_cells('C{0}:D{1}'.format(len(tempData)+2,len(tempData)+2))		
				sk.cell(column=(3),row=(len(tempData)+2),value='Total')
				sk.cell(column=(5),row=(len(tempData)+2),value=ADDDAMA1(tempDataT))
				sk.cell(column=(6),row=(len(tempData)+2),value=ADDDAMA2(tempDataN))
	Tdam = ADDDAMA2(tempDataN)
	tempData.clear()
	tempDataT.clear()
	tempDataN.clear()

	for X,XX in enumerate(wrt):
		for x in XX.keys():
			if XX[x][1][5] == 'OTH':
				tempData[XX[x][1][0]] = XX[x][2]
				tempDataT.append(XX[x][2])
				sk = WB.active = WB['Others']
				for x,xx in enumerate(DEFHID):
					sk.cell(column=(x+1),row=(1),value=DEFHID[x])
	for pa,pA in enumerate(tempData.keys()):
		tempData[pA].insert(0,pa+1)
		for pp,pP in enumerate(tempData[pA]):
			sk.cell(column=(pp+1),row=(pa+2),value=pP)
			sk.cell(column=(5),row=(pa+2),value=tempData[pA][4])
	for X,XX in enumerate(wrt):			
		for x in XX.keys():
			if XX[x][1][5] == 'OTH':
				sk.merge_cells('C{0}:D{1}'.format(len(tempData)+2,len(tempData)+2))		
				sk.cell(column=(3),row=(len(tempData)+2),value='Total')
				sk.cell(column=(5),row=(len(tempData)+2),value=ADD(tempDataT))
	Toth = ADD(tempDataT)
	tempData.clear()
	tempDataT.clear()




	for X,XX in enumerate(wrt):			
		for x in XX.keys():
			tempDataT.append(XX[x][0])
	
	Tven = VenAmtADD(tempDataT)
	Diff = Tven-Tpai-Trfp-Tric-Tpwb-Tnus-Teic-Tdam

	ToDa = datetime.today().strftime('%d-%m-%Y')
	WrittingData = {
		'Lin1': ['Pearson'],
		'Lin2': ['Vendor Reconciliation Statement'],
		'Lin3': ['Company Name:','Pearson','Vendor Name:',''],
		'Lin4': ['Company Code:','','Vendor Code:',''],
		'Lin5': ['Reconciliation as on date:',ToDa,'Reconciliation as on date:',ToDa],
		'Lin6': ['','','Currency',''],
		'Lin7': ['Balance as per vendor statement', VenAmtADD(tempDataT)],
		'Lin8': ['Less'],
		'Lin9': ['Invoices recorded and paid in Oracle (Invoice Paid)',Tpai],
		'Lin10': ['Pending for Payment',Trfp],
		'Lin11': ['Invoice Copy Required', Tric],
		'Lin12': ['Invoice Pending with Pearson Business', Tpwb],
		'Lin13': ['PO Does not Pertain to US', Tnus],
		'Lin14': ['Cancelled Invoice', 0.00],
		'Lin15': ['Pending with Perarson AP', 0.00],
		'Lin16': ['No action required from AP', 0.00],
		'Lin17': ['Pending in COUPA', Teic],
		'Lin18': ['Difference Amount', Tdam],
		'Lin19': ['Duplicate in Vendor Statement', 0.00],
		'Lin20': ['Duplicate Invoice', 0.00],
		'Lin21': ['Add'],
		'Lin22': ['Credit notes in Oracle but not in Vendor Statement', 0.00],
		'Lin23': ['Required Credit Note', 0.00],
		'Lin24': ['Difference', Diff],
		'Lin25': ['', '']
	}


	def WriteSheet(writdat):
		sh = WB.create_sheet('Cons',0)
		sk = WB.active = WB['Cons']
		for da,dda in enumerate(writdat):
			sk.cell(column=(1),row=(da+1),value=writdat[dda][0])
			if da == 3:
				sk.cell(column=(2),row=(da),value=writdat['Lin3'][1])
				sk.cell(column=(7),row=(da),value=writdat['Lin3'][2])
				sk.cell(column=(8),row=(da),value=writdat['Lin3'][3])
			if da == 4:
				sk.cell(column=(2),row=(da),value=writdat['Lin4'][1])
				sk.cell(column=(7),row=(da),value=writdat['Lin4'][2])
				sk.cell(column=(8),row=(da),value=writdat['Lin4'][3])
			if da == 5:
				sk.cell(column=(2),row=(da),value=writdat['Lin5'][1])
				sk.cell(column=(7),row=(da),value=writdat['Lin5'][2])
				sk.cell(column=(8),row=(da),value=writdat['Lin5'][3])
			if da == 6:
				sk.cell(column=(2),row=(da),value=writdat['Lin6'][1])
				sk.cell(column=(7),row=(da),value=writdat['Lin6'][2])
				sk.cell(column=(8),row=(da),value=writdat['Lin6'][3])
			if da == 7:
				sk.cell(column=(1),row=(da),value=writdat['Lin7'][0])
				sk.cell(column=(8),row=(da),value=writdat['Lin7'][1])
				
			if da == 9:
				sk.cell(column=(1),row=(da),value=writdat['Lin9'][0])
				sk.cell(column=(8),row=(da),value=writdat['Lin9'][1])
			if da == 10:
				sk.cell(column=(1),row=(da),value=writdat['Lin10'][0])
				sk.cell(column=(8),row=(da),value=writdat['Lin10'][1])
			if da == 11:
				sk.cell(column=(1),row=(da),value=writdat['Lin11'][0])
				sk.cell(column=(8),row=(da),value=writdat['Lin11'][1])
			if da == 12:
				sk.cell(column=(1),row=(da),value=writdat['Lin12'][0])
				sk.cell(column=(8),row=(da),value=writdat['Lin12'][1])
			if da == 13:
				sk.cell(column=(1),row=(da),value=writdat['Lin13'][0])
				sk.cell(column=(8),row=(da),value=writdat['Lin13'][1])
			if da == 14:
				sk.cell(column=(1),row=(da),value=writdat['Lin14'][0])
				sk.cell(column=(8),row=(da),value=writdat['Lin14'][1])
			if da == 15:
				sk.cell(column=(1),row=(da),value=writdat['Lin15'][0])
				sk.cell(column=(8),row=(da),value=writdat['Lin15'][1])
			if da == 16:
				sk.cell(column=(1),row=(da),value=writdat['Lin16'][0])
				sk.cell(column=(8),row=(da),value=writdat['Lin16'][1])
			if da == 17:
				sk.cell(column=(1),row=(da),value=writdat['Lin17'][0])
				sk.cell(column=(8),row=(da),value=writdat['Lin17'][1])
			if da == 18:
				sk.cell(column=(1),row=(da),value=writdat['Lin18'][0])
				sk.cell(column=(8),row=(da),value=writdat['Lin18'][1])
			if da == 19:
				sk.cell(column=(1),row=(da),value=writdat['Lin19'][0])
				sk.cell(column=(8),row=(da),value=writdat['Lin19'][1])
			if da == 20:
				sk.cell(column=(1),row=(da),value=writdat['Lin20'][0])
				sk.cell(column=(8),row=(da),value=writdat['Lin20'][1])
			if da == 22:
				sk.cell(column=(1),row=(da),value=writdat['Lin22'][0])
				sk.cell(column=(8),row=(da),value=writdat['Lin22'][1])
			if da == 23:
				sk.cell(column=(1),row=(da),value=writdat['Lin23'][0])
				sk.cell(column=(8),row=(da),value=writdat['Lin23'][1])
			if da == 24:
				sk.cell(column=(1),row=(da),value=writdat['Lin24'][0])
				sk.cell(column=(8),row=(da),value=writdat['Lin24'][1])

		sk.merge_cells('A1:I1')
		sk.merge_cells('A2:I2')
		sk.merge_cells('B3:F3')
		sk.merge_cells('H3:I3')
		sk.merge_cells('B4:F4')
		sk.merge_cells('H4:I4')
		sk.merge_cells('B5:F5')
		sk.merge_cells('H5:I5')
		sk.merge_cells('B6:F6')
		sk.merge_cells('H6:I6')
		sk.merge_cells('A7:G7')
		sk.merge_cells('H7:I7')
		sk.merge_cells('A8:G8')
		sk.merge_cells('H8:I8')
		sk.merge_cells('A9:G9')
		sk.merge_cells('H9:I9')
		sk.merge_cells('A10:G10')
		sk.merge_cells('H10:I10')
		sk.merge_cells('A11:G11')
		sk.merge_cells('H11:I11')	
		sk.merge_cells('A12:G12')
		sk.merge_cells('H12:I12')
		sk.merge_cells('A13:G13')
		sk.merge_cells('H13:I13')
		sk.merge_cells('A14:G14')
		sk.merge_cells('H14:I14')
		sk.merge_cells('A15:G15')
		sk.merge_cells('H15:I15')
		sk.merge_cells('A16:G16')
		sk.merge_cells('H16:I16')									
		sk.merge_cells('A17:G17')
		sk.merge_cells('H17:I17')
		sk.merge_cells('A18:G18')
		sk.merge_cells('H18:I18')
		sk.merge_cells('A19:G19')
		sk.merge_cells('H19:I19')
		sk.merge_cells('A20:G20')
		sk.merge_cells('H20:I20')
		sk.merge_cells('A21:G21')
		sk.merge_cells('H21:I21')
		sk.merge_cells('A22:G22')
		sk.merge_cells('H22:I22')
		sk.merge_cells('A23:G23')
		sk.merge_cells('H23:I23')
		sk.merge_cells('A24:G24')
		sk.merge_cells('H24:I24')
		redFill = PatternFill(start_color='c5d9f1',
		                   end_color='c5d9f1',
		                   fill_type='solid')
		pupFill = PatternFill(start_color='e4dfec',
		                   end_color='e4dfec',
		                   fill_type='solid')
		greFill = PatternFill(start_color='d9d9d9',
		                   end_color='d9d9d9',
		                   fill_type='solid')
		yelFill = PatternFill(start_color='ffbf00',
		                   end_color='ffbf00',
		                   fill_type='solid')	                   	                   	                   	
		sk['A1'].fill = redFill
		sk['A2'].fill = redFill
		sk['A3'].fill = pupFill
		sk['B3'].fill = pupFill
		sk['G3'].fill = pupFill
		sk['H3'].fill = pupFill
		sk['A4'].fill = pupFill
		sk['B4'].fill = pupFill
		sk['G4'].fill = pupFill
		sk['H4'].fill = pupFill	
		sk['A5'].fill = pupFill
		sk['B5'].fill = pupFill
		sk['G5'].fill = pupFill
		sk['H5'].fill = pupFill	
		sk['A6'].fill = pupFill
		sk['B6'].fill = pupFill
		sk['G6'].fill = pupFill
		sk['H6'].fill = pupFill
		sk['A8'].fill = greFill
		sk['H8'].fill = greFill
		sk['A21'].fill = greFill
		sk['H21'].fill = greFill	
		sk['A24'].fill = yelFill
		sk['H24'].fill = yelFill

		tempDataT.clear()
	WriteSheet(WrittingData)

WriteTabs(OneData)


def WriteAllData(mrda):
	pass

WB.save("1_Raw_Files\\Output\\ARS_Output_Recon.xlsx")
print('Completed')















# if len(inputFile) <= 0:
# 	print('Input file not found, please recheck and try.')
# else:
# 	pass
